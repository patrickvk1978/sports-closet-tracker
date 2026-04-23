#!/usr/bin/env python3
"""Parse the updated NFL draft workbook into prospects JSON for On the Clock."""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import OrderedDict
from pathlib import Path

import openpyxl


APP_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX_PATH = Path("/Users/patrickvankeerbergen/Downloads/Updated Players.xlsx")
OUTPUT_PATH = APP_DIR / "src" / "data" / "prospects2026.json"

EXPECTED_SHEETS = {
    "ESPN_MOCKDRAFT",
    "ESPN_BIGBOARD",
    "ATHLETIC_mockdraft",
    "ATHLETIC_bigboard",
    "Ringer_MOCKDRAFT",
    "Consensus_MockDraft",
    "Consensus_bigboard",
}

POS_MAP = {
    "CB1": "CB",
    "EDGE1": "EDGE",
    "EDGE/LB": "EDGE",
    "G1": "G",
    "LB1": "LB",
    "OT1": "OT",
    "QB1": "QB",
    "RB1": "RB",
    "S1": "S",
    "TE1": "TE",
    "WR1": "WR",
}


def make_id(name: str) -> str:
    s = unicodedata.normalize("NFD", name.lower())
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[.'`’]", "", s)
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"\s+", "-", s.strip())
    return re.sub(r"-+", "-", s)


def norm_name(name: str) -> str:
    s = unicodedata.normalize("NFD", str(name).strip().lower())
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9\s]", "", s)
    return re.sub(r"\s+", " ", s).strip()


def std_pos(raw: str | None) -> str:
    if raw is None:
        return ""
    value = str(raw).strip().upper()
    value = POS_MAP.get(value, value)
    value = re.sub(r"\d+$", "", value)
    return value.strip()


def parse_name_pos_school(blob: str | None) -> tuple[str, str, str]:
    if not blob:
        return "", "", ""
    parts = [part.strip() for part in str(blob).split(",")]
    name = parts[0] if parts else ""
    position = std_pos(parts[1]) if len(parts) > 1 else ""
    school = ", ".join(parts[2:]).strip() if len(parts) > 2 else ""
    return name, position, school


def load_existing_ids() -> dict[str, str]:
    if not OUTPUT_PATH.exists():
        return {}
    payload = json.loads(OUTPUT_PATH.read_text(encoding="utf-8"))
    return {
        norm_name(item["name"]): item["id"]
        for item in payload.get("prospects", [])
        if item.get("name") and item.get("id")
    }


def parse_espn_bigboard(wb) -> list[dict]:
    ws = wb["ESPN_BIGBOARD"]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        rank = row[0]
        if not isinstance(rank, (int, float)):
            continue
        name, position, school = parse_name_pos_school(row[1])
        if not name:
            continue
        rows.append(
            {
                "rank": int(rank),
                "name": name,
                "position": position,
                "school": school,
            }
        )
    return rows


def parse_athletic_bigboard(wb) -> list[dict]:
    ws = wb["ATHLETIC_bigboard"]
    rows = []
    for rank, name, position, school in ws.iter_rows(min_row=2, values_only=True):
        if not isinstance(rank, (int, float)) or not name:
            continue
        rows.append(
            {
                "rank": int(rank),
                "name": str(name).strip(),
                "position": std_pos(position),
                "school": str(school).strip() if school else "",
            }
        )
    return rows


def parse_consensus_bigboard(wb) -> list[dict]:
    ws = wb["Consensus_bigboard"]
    rows = []
    for rank, name, position, school in ws.iter_rows(min_row=2, values_only=True):
        if not isinstance(rank, (int, float)) or not name:
            continue
        rows.append(
            {
                "rank": int(rank),
                "name": str(name).strip(),
                "position": std_pos(position),
                "school": str(school).strip() if school else "",
            }
        )
    return rows


def parse_espn_mock(wb) -> list[dict]:
    ws = wb["ESPN_MOCKDRAFT"]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        pick = row[0]
        if not isinstance(pick, (int, float)):
            continue
        name, _, _ = parse_name_pos_school(row[2])
        if not name:
            continue
        rows.append({"pick": int(pick), "name": name})
    return rows


def parse_simple_mock_sheet(wb, sheet_name: str) -> list[dict]:
    ws = wb[sheet_name]
    rows = []
    for pick, name, _position, _school in ws.iter_rows(min_row=2, values_only=True):
        if not isinstance(pick, (int, float)) or not name:
            continue
        rows.append({"pick": int(pick), "name": str(name).strip()})
    return rows


def parse_consensus_mock(wb) -> list[dict]:
    ws = wb["Consensus_MockDraft"]
    rows = []
    for pick, name, title in ws.iter_rows(min_row=2, values_only=True):
        if not isinstance(pick, (int, float)) or not name:
            continue
        _name, position, school = parse_name_pos_school(f"{name}, {title}" if title else str(name))
        rows.append(
            {
                "pick": int(pick),
                "name": str(name).strip(),
                "position": position,
                "school": school,
            }
        )
    return rows


def merge_sources(
    espn_bigboard: list[dict],
    athletic_bigboard: list[dict],
    consensus_bigboard: list[dict],
    espn_mock: list[dict],
    athletic_mock: list[dict],
    ringer_mock: list[dict],
    consensus_mock: list[dict],
) -> list[dict]:
    existing_ids = load_existing_ids()
    players: OrderedDict[str, dict] = OrderedDict()

    def get_or_create(name: str, position: str = "", school: str = "") -> dict:
        key = norm_name(name)
        if key not in players:
            players[key] = {
                "id": existing_ids.get(key, make_id(name)),
                "name": name,
                "position": position or "",
                "school": school or "",
                "consensus_rank": None,
                "espn_rank": None,
                "ringer_rank": None,
                "athletic_rank": None,
                "espn_mock_pick": None,
                "ringer_mock_pick": None,
                "athletic_mock_pick": None,
                "consensus_mock_pick": None,
                "predicted_range": None,
                "notes": "",
            }
        player = players[key]
        if position and not player["position"]:
            player["position"] = position
        if school and not player["school"]:
            player["school"] = school
        return player

    for entry in espn_bigboard:
        player = get_or_create(entry["name"], entry["position"], entry["school"])
        player["espn_rank"] = entry["rank"]

    for entry in athletic_bigboard:
        player = get_or_create(entry["name"], entry["position"], entry["school"])
        player["athletic_rank"] = entry["rank"]

    for entry in consensus_bigboard:
        player = get_or_create(entry["name"], entry["position"], entry["school"])
        player["consensus_rank"] = entry["rank"]

    for entry in espn_mock:
        get_or_create(entry["name"])["espn_mock_pick"] = entry["pick"]

    for entry in athletic_mock:
        get_or_create(entry["name"])["athletic_mock_pick"] = entry["pick"]

    for entry in ringer_mock:
        get_or_create(entry["name"])["ringer_mock_pick"] = entry["pick"]

    for entry in consensus_mock:
        player = get_or_create(entry["name"], entry["position"], entry["school"])
        player["consensus_mock_pick"] = entry["pick"]

    prospects = list(players.values())
    for player in prospects:
        mock_picks = [
            player["espn_mock_pick"],
            player["ringer_mock_pick"],
            player["athletic_mock_pick"],
            player["consensus_mock_pick"],
        ]
        picks = [pick for pick in mock_picks if pick is not None]
        if len(picks) == 1:
            pick = picks[0]
            player["predicted_range"] = f"{max(1, pick - 1)}-{pick + 1}"
        elif len(picks) > 1:
            player["predicted_range"] = f"{min(picks)}-{max(picks)}"

    prospects.sort(
        key=lambda item: (
            item["consensus_rank"] if item["consensus_rank"] is not None else 999999,
            item["espn_rank"] if item["espn_rank"] is not None else 999999,
            item["athletic_rank"] if item["athletic_rank"] is not None else 999999,
            item["name"],
        )
    )
    return prospects


def main() -> int:
    xlsx_path = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else DEFAULT_XLSX_PATH
    if not xlsx_path.exists():
        print(f"Workbook not found: {xlsx_path}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    missing = sorted(EXPECTED_SHEETS - set(wb.sheetnames))
    if missing:
        print(f"Missing expected sheets: {', '.join(missing)}", file=sys.stderr)
        return 1

    prospects = merge_sources(
        espn_bigboard=parse_espn_bigboard(wb),
        athletic_bigboard=parse_athletic_bigboard(wb),
        consensus_bigboard=parse_consensus_bigboard(wb),
        espn_mock=parse_espn_mock(wb),
        athletic_mock=parse_simple_mock_sheet(wb, "ATHLETIC_mockdraft"),
        ringer_mock=parse_simple_mock_sheet(wb, "Ringer_MOCKDRAFT"),
        consensus_mock=parse_consensus_mock(wb),
    )

    OUTPUT_PATH.write_text(
        json.dumps({"prospects": prospects}, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    print(f"Wrote {len(prospects)} prospects to {OUTPUT_PATH}")
    print("Top 10:")
    for player in prospects[:10]:
        print(
            f"  {player['consensus_rank'] or '-':>3} "
            f"{player['name']:<28} "
            f"ESPN {player['espn_rank'] or '-':>3} "
            f"ATH {player['athletic_rank'] or '-':>3} "
            f"CON {player['consensus_rank'] or '-':>3} "
            f"EM {player['espn_mock_pick'] or '-':>2} "
            f"RM {player['ringer_mock_pick'] or '-':>2} "
            f"AM {player['athletic_mock_pick'] or '-':>2} "
            f"CM {player['consensus_mock_pick'] or '-':>2}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
